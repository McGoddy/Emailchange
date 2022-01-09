# importing openpyxl and csv
import openpyxl as xl
import csv

exl = 'employeedata.xlsx'

#loading a workbook to contain the excel document
wb = xl.load_workbook(exl)

# Telling python which excel sheet we are working with
sheet = wb['sheet1']

# attributing helpinghands.cm and handsinhand.org
old = 'helpinghands.cm'
new = 'handsinhand.org'

# replacing helpinghands.cm with handsinhand.org
for i in range (2, sheet.max_row + 1):
    cell = sheet.cell(i, 2)
    if old in cell.value:
        updated_email = {cell.value}.replace(old, new)
        sheet.cell(i, 2).value = updated_email
    
# saving the edited excel document to another name to differenciate
wb.save('newemployeedata.xlsx')


# Now we will do same with the .csv file.

# opening and printing out the csv file
op = open('employeedata.csv', 'r')
dt = csv.DictReader(op)
print(dt)
up_dt = []

# replacing helpinghands.cm with handsinhand.org
for r in dt:
    print(r)
    row = {'Employee_Name' : r['Employe_Name'],
           'Email_Address' : r['Email_Address'].replace('helpinghands.cm', 'handsinhand.org'),
           'Phone_Number' : r['Phone_Number'],}
    up_dt.append(row)
print(up_dt)
op.close

# writing the changes to a new .csv file to differenciate 
op = open('newemployeedata.csv', 'w', newline = '')
headers = ['Employee Name', 'Email Address', 'Phone Number']
data = csv.DictWriter(op, delimiter=',', fieldnames= headers)
data.writerow(dict((heads, heads) for heads in headers))
data.writerows(up_dt)
op.close
